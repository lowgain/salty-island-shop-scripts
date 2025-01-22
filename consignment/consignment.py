import datetime
import openpyxl
import csv
import sys
import tkinter as tk
from tkinter.filedialog import askopenfilename
tk.Tk().withdraw() #Discarding unused functions

def getItemList(filename):
  itemList = []
  with open(filename, newline='', errors='ignore') as input:
    data = csv.DictReader(input, delimiter=',')
    for i in data:
      itemList.append(i)
  return filter(lambda x: x['Category']=='Consignment', itemList)

def groupItemsBySupplier(input):
  itemsBySupplier = {}
  for i in input:
      supplier = i['Item name'].split()[0]
      if supplier not in itemsBySupplier:
          itemsBySupplier[supplier] = []
      itemsBySupplier[supplier].append(i)
  return itemsBySupplier

def main():
  wb = openpyxl.Workbook()
  ws = wb.active
  ws['A1'] = 'Total Costs for each person'
  ws.append(['Supplier', 'Total'])
  grandTotal = 0

  filename = askopenfilename(
    title="Select a file",
    filetypes=(("csv files", "*.csv"),("all files", "*.*"))
  )
  itemList = getItemList(filename)
  itemsBySupplier = groupItemsBySupplier(itemList)

  for key, value in itemsBySupplier.items():
    wb.create_sheet(f'{key}')
    ws = wb[f'{key}']
    ws.append(['Name', 'Quantity', 'Rate', 'Ammount'])
    total = 0
    for i in value:
      row = [
        i['Item name'],
        f'{float(i['Items sold']):.0f}',
        f'${float(i['Cost of goods'])/float(i['Items sold']):.2f}',
        f'${float(i['Cost of goods']):.2f}'
      ]
      total += float(i['Cost of goods'])
      ws.append(row)
    grandTotal += total
    ws.append(['','', 'Total:', f'${total:.2f}'])
    ws = wb['Sheet']
    ws.append([f'{key}', f'${total:.2f}'])

  ws = wb['Sheet']
  ws.append(['Grand Total:', f'${grandTotal:.2f}'])
  wb.save(f'consignment-summary-{datetime.datetime.now().date()}.xlsx')

if __name__ == '__main__':
   sys.exit(main())